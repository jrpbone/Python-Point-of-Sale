import csv
import logging
import threading
from pathlib import Path
from tkinter import filedialog, messagebox

from app.ui.dialogs import ErrorListDialog, ImportPreviewDialog, ProgressDialog

REQUIRED_HEADERS = ["name", "sku", "category", "price", "quantity", "brand"]
OPTIONAL_HEADERS = ["barcode"]
PREVIEW_ROWS = 10
PROGRESS_THRESHOLD = 200
PROGRESS_UPDATE_STEP = 5


def _schedule(parent, func, *args):
    if parent is None:
        func(*args)
        return
    parent.after(0, lambda: func(*args))


def _create_worker_service():
    from app.db import get_connection
    from app.services.pos_service import PosService

    return PosService(get_connection())


def load_products_from_excel(
    service,
    default_path=None,
    prompt_if_missing=True,
    parent=None,
    force_prompt=False,
    admin_user_id=None,
    on_complete=None,
):
    path = None
    if default_path and Path(default_path).exists() and not force_prompt:
        path = str(default_path)
    if not path and prompt_if_missing:
        path = filedialog.askopenfilename(
            title="Select products file",
            filetypes=[("Excel or CSV Files", "*.xlsx;*.csv"), ("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")],
            initialdir=str(Path(default_path).parent) if default_path else None,
            parent=parent,
        )
    if not path:
        if on_complete:
            on_complete(False, None)
        return False

    rows, headers = _read_product_rows(path, parent)
    if rows is None:
        if on_complete:
            on_complete(False, None)
        return False
    if not rows:
        messagebox.showerror("Import failed", "No data rows found in the file.", parent=parent)
        if on_complete:
            on_complete(False, None)
        return False

    try:
        preview = service.preview_import(rows)
    except Exception as exc:
        messagebox.showerror("Import failed", str(exc), parent=parent)
        if on_complete:
            on_complete(False, None)
        return False
    valid_rows = preview.get("valid_rows", [])
    errors = preview.get("errors", [])
    summary = preview.get("summary", {})
    if not valid_rows:
        messagebox.showerror("Import failed", "No valid rows to import.", parent=parent)
        if errors:
            ErrorListDialog(parent, "Import Errors", errors)
        if on_complete:
            on_complete(False, None)
        return False

    preview_columns = [
        ("name", "Name"),
        ("sku", "SKU"),
        ("category", "Category"),
        ("price", "Price"),
        ("quantity", "Quantity"),
        ("brand", "Brand"),
        ("__action", "Action"),
    ]
    preview_rows = valid_rows[:PREVIEW_ROWS]
    dialog = ImportPreviewDialog(
        parent,
        columns=preview_columns,
        rows=preview_rows,
        summary=summary,
        errors=errors,
    )
    parent.wait_window(dialog.top)
    if not dialog.result:
        if on_complete:
            on_complete(False, None)
        return False

    progress = None
    if summary.get("valid", 0) >= PROGRESS_THRESHOLD:
        progress = ProgressDialog(parent, "Importing products", summary.get("valid", 0))

    def _progress_callback(current, total):
        if not progress:
            return
        if current % PROGRESS_UPDATE_STEP == 0 or current == total:
            _schedule(parent, progress.update_progress, current, total)

    def _finish_import(stats=None, error=None):
        if progress:
            progress.close()
        if error:
            messagebox.showerror("Import failed", error, parent=parent)
            if on_complete:
                on_complete(False, None)
            return
        added = stats.get("added")
        updated = stats.get("updated")
        details = ""
        if added is not None and updated is not None:
            details = f"\nAdded {added} new products, updated {updated} existing."
        messagebox.showinfo(
            "Import complete",
            f"Loaded {len(valid_rows)} products from file.{details}",
            parent=parent,
        )

        all_errors = list(errors)
        all_errors.extend(stats.get("errors", []))
        if all_errors:
            ErrorListDialog(parent, "Import Errors", all_errors)
        if on_complete:
            on_complete(True, stats)

    def _worker():
        try:
            worker_service = _create_worker_service()
        except Exception as exc:
            logging.exception("Import worker setup failed")
            _schedule(parent, _finish_import, None, str(exc))
            return
        try:
            stats = worker_service.import_products(
                valid_rows,
                audit_user_id=admin_user_id,
                source_label=Path(path).name,
                progress_callback=_progress_callback,
            ) or {}
            _schedule(parent, _finish_import, stats, None)
        except Exception as exc:
            logging.exception("Import failed")
            _schedule(parent, _finish_import, None, str(exc))
        finally:
            try:
                worker_service.conn.close()
            except Exception:
                pass

    threading.Thread(target=_worker, daemon=True).start()
    return True


def export_products_to_excel(service, output_path, parent=None, on_complete=None):
    if not output_path:
        messagebox.showerror(
            "Export failed",
            "No export path selected.",
            parent=parent,
        )
        if on_complete:
            on_complete({"success": False, "error": "No export path selected."})
        return False
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        messagebox.showerror(
            "Missing dependency",
            "openpyxl is required to write .xlsx files. Install it and try again.",
            parent=parent,
        )
        if on_complete:
            on_complete({"success": False, "error": "openpyxl is required to write .xlsx files."})
        return False

    output_path = Path(output_path)
    progress = ProgressDialog(parent, "Exporting products", 1)

    def _finish_export(result):
        progress.close()
        if result.get("success"):
            messagebox.showinfo(
                "Export complete",
                f"Exported {result.get('count', 0)} products to {output_path.name}.",
                parent=parent,
            )
        else:
            messagebox.showerror("Export failed", result.get("error", "Unknown error"), parent=parent)
        if on_complete:
            on_complete(result)

    def _worker():
        try:
            worker_service = _create_worker_service()
        except Exception as exc:
            logging.exception("Export worker setup failed")
            _schedule(parent, _finish_export, {"success": False, "error": str(exc)})
            return
        try:
            result = _export_products_worker(worker_service, output_path)
            _schedule(parent, _finish_export, result)
        except Exception as exc:
            logging.exception("Export failed")
            _schedule(parent, _finish_export, {"success": False, "error": str(exc)})
        finally:
            try:
                worker_service.conn.close()
            except Exception:
                pass

    threading.Thread(target=_worker, daemon=True).start()
    return True


def _export_products_worker(service, output_path):
    from openpyxl import Workbook

    products = service.list_products()
    headers = ["Name", "SKU", "Category", "Price", "Quantity", "Brand", "Barcode"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(headers)
    for product in products:
        sheet.append(
            [
                product.name,
                product.sku or "",
                product.category or "",
                float(product.price),
                int(product.quantity),
                product.brand or "",
                product.barcode or "",
            ]
        )
    workbook.save(output_path)
    return {"success": True, "count": len(products), "path": output_path}


def _read_product_rows(path, parent):
    file_path = Path(path)
    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        return _read_excel_rows(file_path, parent)
    if suffix == ".csv":
        return _read_csv_rows(file_path, parent)
    messagebox.showerror("Invalid file", "Select a .xlsx or .csv file to import.", parent=parent)
    return None, None


def _read_excel_rows(path, parent):
    try:
        from openpyxl import load_workbook
    except ImportError:
        messagebox.showerror(
            "Missing dependency",
            "openpyxl is required to read .xlsx files. Install it and try again.",
            parent=parent,
        )
        return None, None
    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        headers = _normalize_headers([cell.value for cell in sheet[1]])
        if not _has_required_headers(headers):
            messagebox.showerror(
                "Invalid file",
                "File must include columns: Name, SKU, Category, Price, Quantity, Brand. Barcode is optional.",
                parent=parent,
            )
            return None, None
        rows = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            record = {
                headers[col_idx]: (value if value is not None else "")
                for col_idx, value in enumerate(row)
                if col_idx < len(headers)
            }
            record["_row_number"] = idx
            rows.append(record)
        return rows, headers
    except Exception as exc:
        messagebox.showerror("Import failed", str(exc), parent=parent)
        return None, None


def _read_csv_rows(path, parent):
    try:
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                messagebox.showerror("Invalid file", "CSV file is missing a header row.", parent=parent)
                return None, None
            header_map = _normalize_header_map(reader.fieldnames)
            headers = list(header_map.values())
            if not _has_required_headers(headers):
                messagebox.showerror(
                    "Invalid file",
                    "CSV must include columns: Name, SKU, Category, Price, Quantity, Brand. Barcode is optional.",
                    parent=parent,
                )
                return None, None
            rows = []
            for idx, row in enumerate(reader, start=2):
                if not any((value or "").strip() for value in row.values()):
                    continue
                record = {
                    header_map[key]: (value if value is not None else "")
                    for key, value in row.items()
                    if key in header_map
                }
                record["_row_number"] = idx
                rows.append(record)
            return rows, headers
    except Exception as exc:
        messagebox.showerror("Import failed", str(exc), parent=parent)
        return None, None


def _normalize_headers(headers):
    return [str(header or "").strip().lower() for header in headers]


def _normalize_header_map(headers):
    header_map = {}
    for header in headers:
        normalized = str(header or "").strip().lower()
        if not normalized:
            continue
        header_map[header] = normalized
    return header_map


def _has_required_headers(headers):
    return all(field in headers for field in REQUIRED_HEADERS)
