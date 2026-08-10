import sqlite3
import sys
from datetime import datetime
from pathlib import Path

from app.settings import TAX_RATE, TAX_ROUNDING
from app.db import DB_PATH
from app.repositories.audit_repository import AuditRepository
from app.repositories.category_repository import CategoryRepository
from app.repositories.payment_repository import PaymentRepository
from app.repositories.product_repository import ProductRepository
from app.repositories.sale_repository import SaleRepository
from app.repositories.stock_repository import StockRepository
from app.repositories.user_repository import UserRepository


class PosService:
    MAX_IMPORT_PRICE = 1_000_000.0
    MAX_IMPORT_QTY = 1_000_000

    def __init__(self, conn):
        self.conn = conn
        self.products = ProductRepository(conn)
        self.categories = CategoryRepository(conn)
        self.users = UserRepository(conn)
        self.sales = SaleRepository(conn)
        self.payments = PaymentRepository(conn)
        self.stock = StockRepository(conn)
        self.audit = AuditRepository(conn)
        # Frozen one-file builds run from a temporary extraction directory.
        # Keep generated exports beside the persistent application database.
        self.base_dir = (
            DB_PATH.parent
            if getattr(sys, "frozen", False)
            else Path(__file__).resolve().parents[2]
        )

    def reconnect(self):
        try:
            self.conn.close()
        except Exception:
            pass
        from app.db import get_connection

        self.reset_connection(get_connection())

    def _ensure_connection(self):
        try:
            self.conn.execute("SELECT 1")
        except sqlite3.Error:
            self.reconnect()

    def login(self, username, pin):
        self._ensure_connection()
        return self.users.authenticate(username, pin)

    def authorize_admin_pin(self, pin):
        self._ensure_connection()
        return self.users.authenticate_admin(pin)

    def reset_connection(self, conn):
        self.conn = conn
        self.products = ProductRepository(conn)
        self.categories = CategoryRepository(conn)
        self.users = UserRepository(conn)
        self.sales = SaleRepository(conn)
        self.payments = PaymentRepository(conn)
        self.stock = StockRepository(conn)
        self.audit = AuditRepository(conn)

    def list_products(self):
        self._ensure_connection()
        return self.products.list_products()

    def list_users(self, active=None):
        self._ensure_connection()
        return self.users.list_users(active)

    def create_user(self, username, pin, first_name=None, role="cashier"):
        self._ensure_connection()
        try:
            user = self.users.create_user(username, pin, first_name, role)
            self.conn.commit()
            return user
        except Exception:
            self.conn.rollback()
            raise

    def update_username(self, user_id, new_username):
        self._ensure_connection()
        try:
            self.users.update_username(user_id, new_username)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def reset_user_pin(self, user_id, new_pin):
        self._ensure_connection()
        try:
            self.users.reset_pin(user_id, new_pin)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def set_user_active(self, user_id, active):
        self._ensure_connection()
        try:
            row = self.users.get_user_by_id(user_id)
            if not row:
                raise ValueError("User not found.")
            if not active and row["role"] != "cashier":
                raise ValueError("Only cashier users can be made inactive.")
            self.users.set_active(user_id, active)
            self.conn.commit()
            return self.users.get_user_by_id(user_id)
        except Exception:
            self.conn.rollback()
            raise

    def get_product_by_sku(self, sku):
        self._ensure_connection()
        return self.products.get_by_sku(sku)

    def preview_import(self, rows):
        self._ensure_connection()
        errors = []
        valid_rows = []
        summary = {"total": len(rows), "valid": 0, "invalid": 0, "add": 0, "update": 0}
        skus = []
        for row in rows:
            sku = self._clean_text(row.get("sku")).strip()
            if sku:
                skus.append(sku)
        existing_map = self.products.get_sku_map(list(dict.fromkeys(skus)))
        for idx, row in enumerate(rows, start=1):
            row_number = row.get("_row_number", idx)
            sku = self._clean_text(row.get("sku"))
            sku = sku.strip()
            if not sku:
                errors.append(f"Row {row_number}: Missing SKU.")
                continue

            existing = existing_map.get(sku)
            quantity = self._parse_int(row.get("quantity"), default=0)
            if quantity is None or quantity < 0 or quantity > self.MAX_IMPORT_QTY:
                errors.append(
                    f"Row {row_number} (SKU {sku}): Invalid quantity '{row.get('quantity')}'."
                )
                continue

            if existing:
                cleaned = {
                    "name": self._clean_text(row.get("name")),
                    "sku": sku,
                    "barcode": self._clean_text(row.get("barcode")) or None,
                    "category": self._clean_text(row.get("category")),
                    "price": self._parse_float(row.get("price")),
                    "quantity": quantity,
                    "brand": self._clean_text(row.get("brand")) or None,
                    "_row_number": row_number,
                    "__action": "update",
                    "_existing_id": existing["id"],
                    "_before_qty": existing["quantity"],
                }
                valid_rows.append(cleaned)
                summary["update"] += 1
                continue

            name = self._clean_text(row.get("name"))
            if not name:
                errors.append(f"Row {row_number} (SKU {sku}): Missing name.")
                continue
            price = self._parse_float(row.get("price"))
            if price is None or price < 0 or price > self.MAX_IMPORT_PRICE:
                errors.append(
                    f"Row {row_number} (SKU {sku}): Invalid price '{row.get('price')}'."
                )
                continue
            cleaned = {
                "name": name,
                "sku": sku,
                "barcode": self._clean_text(row.get("barcode")) or None,
                "category": self._clean_text(row.get("category")),
                "price": price,
                "quantity": quantity,
                "brand": self._clean_text(row.get("brand")) or None,
                "_row_number": row_number,
                "__action": "add",
                "_before_qty": 0,
            }
            valid_rows.append(cleaned)
            summary["add"] += 1

        summary["valid"] = len(valid_rows)
        summary["invalid"] = summary["total"] - summary["valid"]
        return {"valid_rows": valid_rows, "errors": errors, "summary": summary}

    def import_products(
        self,
        rows,
        audit_user_id=None,
        source_label=None,
        progress_callback=None,
        skip_invalids=True,
        dry_run=False,
    ):
        self._ensure_connection()
        added = 0
        updated = 0
        errors = []
        total = len(rows)
        audit_entries = []

        skus = []
        for row in rows:
            sku = self._clean_text(row.get("sku")).strip()
            if sku:
                skus.append(sku)
        existing_map = self.products.get_sku_map(list(dict.fromkeys(skus)))
        sku_state = {
            sku: {"id": row["id"], "quantity": row["quantity"]}
            for sku, row in existing_map.items()
        }
        insert_data = {}

        for idx, row in enumerate(rows, start=1):
            try:
                sku = self._clean_text(row.get("sku")).strip()
                if not sku:
                    errors.append(f"Row {row.get('_row_number', idx)}: Missing SKU.")
                    continue

                quantity = row.get("quantity")
                if not isinstance(quantity, int):
                    quantity = self._parse_int(quantity, default=0)
                if quantity is None or quantity < 0 or quantity > self.MAX_IMPORT_QTY:
                    errors.append(
                        f"Row {row.get('_row_number', idx)} (SKU {sku}): Invalid quantity '{row.get('quantity')}'."
                    )
                    continue

                is_existing = sku in sku_state
                if not is_existing:
                    name = self._clean_text(row.get("name"))
                    if not name:
                        errors.append(
                            f"Row {row.get('_row_number', idx)} (SKU {sku}): Missing name."
                        )
                        continue
                    price = row.get("price")
                    if not isinstance(price, (float, int)):
                        price = self._parse_float(price)
                    if price is None or price < 0 or price > self.MAX_IMPORT_PRICE:
                        errors.append(
                            f"Row {row.get('_row_number', idx)} (SKU {sku}): Invalid price '{row.get('price')}'."
                        )
                        continue

                    insert_data[sku] = {
                        "name": name,
                        "sku": sku,
                        "barcode": self._clean_text(row.get("barcode")) or None,
                        "category": self._clean_text(row.get("category")),
                        "price": float(price),
                        "quantity": 0,
                        "brand": self._clean_text(row.get("brand")) or None,
                    }
                    sku_state[sku] = {"id": None, "quantity": 0}
                    added += 1
                else:
                    updated += 1

                before_qty = sku_state[sku]["quantity"]
                after_qty = before_qty + quantity
                sku_state[sku]["quantity"] = after_qty
                if sku in insert_data:
                    insert_data[sku]["quantity"] = after_qty

                if audit_user_id:
                    details = f"SKU {sku} qty {before_qty} -> {after_qty} (delta {quantity})"
                    if source_label:
                        details = f"{details} from {source_label}"
                    audit_entries.append((audit_user_id, "IMPORT_PRODUCT", details))
            except Exception as exc:
                errors.append(
                    f"Row {row.get('_row_number', idx)} (SKU {row.get('sku')}): {exc}"
                )
            finally:
                if progress_callback:
                    try:
                        progress_callback(idx, total)
                    except Exception:
                        pass

        if errors and not skip_invalids:
            return {"added": 0, "updated": 0, "errors": errors, "aborted": True}

        if dry_run:
            return {"added": added, "updated": updated, "errors": errors}

        updates = []
        for sku, state in sku_state.items():
            if state["id"] is None:
                continue
            if sku not in existing_map:
                continue
            before_qty = existing_map[sku]["quantity"]
            new_qty = state["quantity"]
            if new_qty != before_qty:
                updates.append((new_qty, state["id"]))

        inserts = []
        category_cache = {}

        try:
            self.conn.execute("BEGIN")
            for data in insert_data.values():
                category_name = data["category"]
                if not category_name:
                    category_id = None
                elif category_name in category_cache:
                    category_id = category_cache[category_name]
                else:
                    category_id = self.categories.get_or_create(category_name)
                    category_cache[category_name] = category_id
                inserts.append(
                    (
                        data["name"],
                        data["sku"],
                        data["barcode"],
                        category_id,
                        data["price"],
                        data["quantity"],
                        data["brand"],
                    )
                )
            self.products.update_quantities_bulk(updates)
            self.products.insert_many(inserts)
            if audit_entries:
                self.audit.add_entries(audit_entries)
            if audit_user_id:
                summary = (
                    f"Imported products from {source_label or 'file'}; "
                    f"added {added}, updated {updated}, errors {len(errors)}"
                )
                self.audit.add_entry(audit_user_id, "IMPORT_PRODUCTS", summary)
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

        return {"added": added, "updated": updated, "errors": errors}

    def log_admin_action(self, user_id, action, details=None):
        self._ensure_connection()
        try:
            self.audit.add_entry(user_id, action, details)
            self.conn.commit()
            return True
        except sqlite3.ProgrammingError:
            try:
                from app.db import get_connection
            except Exception:
                return False
            try:
                new_conn = get_connection()
                self.reset_connection(new_conn)
                self.audit.add_entry(user_id, action, details)
                self.conn.commit()
                return True
            except Exception:
                return False
        except Exception:
            return False

    @staticmethod
    def _clean_text(value):
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _parse_float(value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace(",", "")
        if text[:1] in ("P", "p", "$"):
            text = text[1:]
        try:
            return float(text)
        except ValueError:
            return None

    @classmethod
    def _parse_int(cls, value, default=None):
        if value is None or (isinstance(value, str) and not value.strip()):
            return default
        parsed = cls._parse_float(value)
        if parsed is None:
            return None
        if parsed % 1 != 0:
            return None
        return int(parsed)

    def get_sales_summary(self):
        self._ensure_connection()
        periods = {
            "today": "datetime('now','localtime','start of day')",
            "week": "datetime('now','localtime','-6 days','start of day')",
            "month": "datetime('now','localtime','-29 days','start of day')",
        }
        totals = {key: self.sales.totals_since(expr) for key, expr in periods.items()}
        top_skus = self.sales.top_skus_since(periods["month"], limit=5)
        return {"totals": totals, "top_skus": top_skus}

    def checkout(self, user_id, cart_items, payment_method, received_amount, discount=0.0):
        self._ensure_connection()
        if not cart_items:
            raise ValueError("Cart is empty")
        if discount < 0:
            raise ValueError("Discount cannot be negative")

        subtotal = sum(item.line_total for item in cart_items)
        taxable = max(subtotal - discount, 0)
        tax = round(taxable * TAX_RATE, TAX_ROUNDING)
        total = round(taxable + tax, TAX_ROUNDING)

        if payment_method == "CASH" and received_amount < total:
            raise ValueError("Insufficient cash received")

        try:
            self.conn.execute("BEGIN")

            for item in cart_items:
                row = self.products.get_by_id(item.product_id)
                if not row:
                    raise ValueError("Product not found")
                if row["quantity"] < item.quantity:
                    raise ValueError(f"Not enough stock for {row['name']}")

            sale_id = self.sales.create_sale(
                user_id=user_id,
                subtotal=subtotal,
                discount=discount,
                tax=tax,
                total=total,
            )

            for item in cart_items:
                self.sales.add_line_item(
                    sale_id,
                    item.product_id,
                    item.quantity,
                    item.unit_price,
                    item.line_total,
                )
                row = self.products.get_by_id(item.product_id)
                new_qty = row["quantity"] - item.quantity
                self.products.update_quantity(item.product_id, new_qty)
                self.stock.add_movement(
                    item.product_id,
                    -item.quantity,
                    "OUT",
                    f"Sale #{sale_id}",
                )

            change_due = round(received_amount - total, 2)
            self.payments.add_payment(
                sale_id,
                payment_method,
                total,
                received_amount,
                change_due,
            )
            self.audit.add_entry(user_id, "SALE", f"Sale #{sale_id} total {total}")

            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

        self._export_sale(
            sale_id=sale_id,
            user_id=user_id,
            subtotal=subtotal,
            discount=discount,
            total=total,
            payment_method=payment_method,
            received=received_amount,
            change_due=change_due,
            cart_items=cart_items,
        )

        return {
            "sale_id": sale_id,
            "subtotal": subtotal,
            "discount": discount,
            "tax": tax,
            "total": total,
            "change_due": change_due,
        }

    def _export_sale(self, sale_id, user_id, subtotal, discount, total, payment_method, received, change_due, cart_items):
        timestamp = datetime.now().isoformat(timespec="seconds")
        items = "; ".join(
            f"{item.name} x{item.quantity} @ {item.unit_price:.2f}"
            for item in cart_items
        )
        record = {
            "sale_id": sale_id,
            "timestamp": timestamp,
            "user_id": user_id,
            "subtotal": f"{subtotal:.2f}",
            "discount": f"{discount:.2f}",
            "total": f"{total:.2f}",
            "payment_method": payment_method,
            "amount_received": f"{received:.2f}",
            "change_due": f"{change_due:.2f}",
            "items": items,
        }
        headers = list(record.keys())
        xlsx_path = self.base_dir / "sales.xlsx"

        try:
            from openpyxl import Workbook, load_workbook
        except ImportError as exc:
            raise ValueError("openpyxl is required to export sales.xlsx") from exc

        if xlsx_path.exists():
            workbook = load_workbook(xlsx_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sales"
            sheet.append(headers)
        sheet.append([record[key] for key in headers])
        workbook.save(xlsx_path)
