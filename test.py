"""Comprehensive test suite for PyPOS.

Run from the repository root with::

    python test.py

All database and spreadsheet operations use a temporary directory. The real
application database and exports are never modified.
"""

import sqlite3
import tempfile
import tkinter as tk
import unittest
from pathlib import Path
from tkinter import ttk
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import app.db as db
from app.security import hash_pin, is_hashed, verify_pin
from app.services.pos_service import PosService
from app.ui.config import APP_ICON_PATH, apply_theme, apply_window_icon, get_palette
from app.ui.controllers.pos_controller import PosController
from app.ui.dialogs.admin_dialog import AdminAuthDialog
from app.ui.dialogs.import_dialogs import (
    ErrorListDialog,
    ImportPreviewDialog,
    ProgressDialog,
)
from app.ui.dialogs.manage_users_dialog import ManageUsersDialog
from app.ui.dialogs.payment_dialog import PaymentDialog
from app.ui.dialogs.report_dialog import ReportDialog
from app.ui.dialogs.restore_dialog import RestoreBackupDialog
from app.ui.formatting import format_currency
from app.ui import importers
from app.ui.services.cart_service import CartService
from app.ui.state import MainWindowState
from app.ui.view_models import CartRow
from app.ui.views.login_view import LoginView
from app.ui.views.pos_view import PosView


class DatabaseTestCase(unittest.TestCase):
    """Base class that provides a fully initialized temporary application DB."""

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.temp_dir.name)
        self.original_db_path = db.DB_PATH
        db.DB_PATH = self.temp_path / "pos.db"
        db.init_db()
        db.seed_data()
        self.conn = db.get_connection()
        self.service = PosService(self.conn)
        self.service.base_dir = self.temp_path
        self.admin = self.service.login("admin", "admin")

    def tearDown(self):
        try:
            self.conn.close()
        except Exception:
            pass
        db.DB_PATH = self.original_db_path
        self.temp_dir.cleanup()

    def add_product(
        self,
        sku="SKU-001",
        name="Test Product",
        price=25.0,
        quantity=10,
        barcode="BAR-001",
        category="General",
        brand="PyPOS",
    ):
        result = self.service.import_products(
            [
                {
                    "name": name,
                    "sku": sku,
                    "barcode": barcode,
                    "category": category,
                    "price": price,
                    "quantity": quantity,
                    "brand": brand,
                }
            ]
        )
        self.assertEqual(result["errors"], [])
        return self.service.get_product_by_sku(sku)


class SecurityTests(unittest.TestCase):
    def test_pin_hashing_and_verification(self):
        first = hash_pin("1234", iterations=1_000)
        second = hash_pin("1234", iterations=1_000)
        self.assertNotEqual(first, second)
        self.assertTrue(is_hashed(first))
        self.assertTrue(verify_pin("1234", first))
        self.assertFalse(verify_pin("9999", first))

    def test_invalid_hashes_and_missing_pin(self):
        self.assertFalse(is_hashed("plain-text"))
        self.assertFalse(verify_pin("1234", "invalid$hash"))
        self.assertFalse(verify_pin("1234", "pbkdf2_sha256$bad$x$y"))
        with self.assertRaises(ValueError):
            hash_pin(None)


class DatabaseAndBackupTests(DatabaseTestCase):
    def test_schema_seed_and_foreign_keys(self):
        expected_tables = {
            "categories",
            "products",
            "users",
            "sales",
            "sale_items",
            "payments",
            "stock_movements",
            "audit_logs",
        }
        tables = {
            row["name"]
            for row in self.conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
        self.assertTrue(expected_tables.issubset(tables))
        self.assertEqual(self.conn.execute("PRAGMA foreign_keys").fetchone()[0], 1)
        self.assertIsNotNone(self.admin)
        self.assertEqual(self.admin.role, "admin")

        stored = self.conn.execute(
            "SELECT pin FROM users WHERE username = 'admin'"
        ).fetchone()["pin"]
        self.assertTrue(is_hashed(stored))

        db.seed_data()
        count = self.conn.execute(
            "SELECT COUNT(*) AS count FROM users WHERE username = 'admin'"
        ).fetchone()["count"]
        self.assertEqual(count, 1)

    def test_backup_restore_and_pre_restore_safety_copy(self):
        backup = db.backup_database()
        self.assertTrue(backup.exists())

        self.conn.execute("INSERT INTO categories (name) VALUES ('Temporary')")
        self.conn.commit()
        self.conn.close()

        safety = db.restore_database(backup, create_pre_restore_backup=True)
        self.assertIsNotNone(safety)
        self.assertTrue(safety.exists())

        self.conn = db.get_connection()
        restored = self.conn.execute(
            "SELECT id FROM categories WHERE name = 'Temporary'"
        ).fetchone()
        self.assertIsNone(restored)

    def test_backup_and_restore_reject_missing_files(self):
        missing = self.temp_path / "missing.db"
        with self.assertRaises(FileNotFoundError):
            db.backup_database(source_path=missing)
        with self.assertRaises(FileNotFoundError):
            db.restore_database(missing)


class UserManagementTests(DatabaseTestCase):
    def test_user_lifecycle_and_authentication(self):
        cashier = self.service.create_user("cashier1", "2468", "Maria", "cashier")
        self.assertEqual(cashier["role"], "cashier")
        self.assertEqual(self.service.login("cashier1", "2468").first_name, "Maria")
        self.assertEqual(self.service.login("Maria", "2468").username, "cashier1")
        self.assertIsNone(self.service.login("cashier1", "wrong"))

        self.service.update_username(cashier["id"], "cashier2")
        self.assertIsNone(self.service.login("cashier1", "2468"))
        self.assertIsNotNone(self.service.login("cashier2", "2468"))

        self.service.reset_user_pin(cashier["id"], "1357")
        self.assertIsNone(self.service.login("cashier2", "2468"))
        self.assertIsNotNone(self.service.login("cashier2", "1357"))

        updated = self.service.set_user_active(cashier["id"], False)
        self.assertEqual(updated["active"], 0)
        self.assertIsNone(self.service.login("cashier2", "1357"))
        self.assertEqual(len(self.service.list_users(active=False)), 1)

        self.service.set_user_active(cashier["id"], True)
        self.assertIsNotNone(self.service.login("cashier2", "1357"))

    def test_user_validation_and_admin_protection(self):
        for username, pin, role in (
            ("", "1234", "cashier"),
            ("user", "", "cashier"),
            ("user", "1234", "owner"),
        ):
            with self.assertRaises(ValueError):
                self.service.create_user(username, pin, role=role)

        with self.assertRaises(ValueError):
            self.service.set_user_active(self.admin.id, False)
        with self.assertRaises(ValueError):
            self.service.set_user_active(999_999, False)

    def test_plain_text_legacy_pin_is_migrated_after_login(self):
        cursor = self.conn.execute(
            "INSERT INTO users (username, pin, role) VALUES (?, ?, ?)",
            ("legacy", "0000", "cashier"),
        )
        self.conn.commit()
        self.assertIsNotNone(self.service.login("legacy", "0000"))
        stored = self.conn.execute(
            "SELECT pin FROM users WHERE id = ?", (cursor.lastrowid,)
        ).fetchone()["pin"]
        self.assertTrue(is_hashed(stored))


class ProductAndImportTests(DatabaseTestCase):
    def test_preview_classifies_valid_and_invalid_rows(self):
        self.add_product(sku="EXISTING", quantity=4)
        preview = self.service.preview_import(
            [
                {
                    "name": "Existing",
                    "sku": "EXISTING",
                    "price": "ignored for update",
                    "quantity": "2",
                },
                {
                    "name": "New",
                    "sku": "NEW",
                    "price": "P1,250.50",
                    "quantity": "3",
                    "category": "Retail",
                },
                {"name": "Missing SKU", "sku": "", "price": 1, "quantity": 1},
                {"name": "Bad Qty", "sku": "BAD-QTY", "price": 1, "quantity": 1.5},
                {"name": "Bad Price", "sku": "BAD-PRICE", "price": "x", "quantity": 1},
            ]
        )
        self.assertEqual(preview["summary"], {
            "total": 5,
            "valid": 2,
            "invalid": 3,
            "add": 1,
            "update": 1,
        })
        self.assertEqual(len(preview["errors"]), 3)
        self.assertEqual(preview["valid_rows"][0]["__action"], "update")
        self.assertEqual(preview["valid_rows"][1]["price"], 1250.50)

    def test_import_add_update_barcode_lookup_and_audit(self):
        progress = []
        rows = [
            {
                "name": "Coffee",
                "sku": "COF-1",
                "barcode": "4800001",
                "category": "Drinks",
                "price": 80,
                "quantity": 5,
                "brand": "House",
            },
            {
                "name": "Coffee",
                "sku": "COF-1",
                "barcode": "4800001",
                "category": "Drinks",
                "price": 80,
                "quantity": 2,
                "brand": "House",
            },
        ]
        result = self.service.import_products(
            rows,
            audit_user_id=self.admin.id,
            source_label="products.xlsx",
            progress_callback=lambda current, total: progress.append((current, total)),
        )
        self.assertEqual(result["added"], 1)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(progress, [(1, 2), (2, 2)])
        product = self.service.get_product_by_sku("COF-1")
        self.assertEqual(product["quantity"], 7)
        self.assertEqual(self.service.get_product_by_sku("4800001")["id"], product["id"])

        products = self.service.list_products()
        coffee = next(item for item in products if item.sku == "COF-1")
        self.assertEqual(coffee.category, "Drinks")
        self.assertEqual(coffee.brand, "House")

        actions = [
            row["action"]
            for row in self.conn.execute(
                "SELECT action FROM audit_logs ORDER BY id"
            ).fetchall()
        ]
        self.assertIn("IMPORT_PRODUCT", actions)
        self.assertIn("IMPORT_PRODUCTS", actions)

    def test_strict_import_abort_and_dry_run_do_not_write(self):
        rows = [
            {"name": "Valid", "sku": "VALID", "price": 10, "quantity": 1},
            {"name": "Invalid", "sku": "", "price": 10, "quantity": 1},
        ]
        aborted = self.service.import_products(rows, skip_invalids=False)
        self.assertTrue(aborted["aborted"])
        self.assertIsNone(self.service.get_product_by_sku("VALID"))

        dry_run = self.service.import_products(rows[:1], dry_run=True)
        self.assertEqual(dry_run["added"], 1)
        self.assertIsNone(self.service.get_product_by_sku("VALID"))

    def test_parse_helpers_cover_currency_and_integer_rules(self):
        self.assertEqual(self.service._parse_float("P1,234.50"), 1234.5)
        self.assertEqual(self.service._parse_float("$5"), 5.0)
        self.assertIsNone(self.service._parse_float("not money"))
        self.assertEqual(self.service._parse_int("4.0"), 4)
        self.assertIsNone(self.service._parse_int("4.5"))
        self.assertEqual(self.service._parse_int("", default=7), 7)


class CartTests(DatabaseTestCase):
    def test_cart_add_increment_remove_clear_and_stock_limits(self):
        product = self.add_product(price=12.5, quantity=4)
        cart = {}
        service = CartService(cart)

        service.add_product(product, 2)
        service.add_product(product, 1)
        self.assertTrue(service.has_items())
        self.assertEqual(len(service.checkout_items()), 1)
        self.assertEqual(service.checkout_items()[0].quantity, 3)
        self.assertEqual(service.subtotal(), 37.5)

        with self.assertRaises(ValueError):
            service.add_product(product, 2)

        self.assertTrue(service.remove_one(product["id"]))
        self.assertEqual(service.checkout_items()[0].quantity, 2)
        self.assertFalse(service.remove_one(999_999))
        service.clear()
        self.assertFalse(service.has_items())


class CheckoutAndReportingTests(DatabaseTestCase):
    def cart_row(self, product, quantity):
        return CartRow(
            product_id=product["id"],
            name=product["name"],
            sku=product["sku"],
            unit_price=product["price"],
            quantity=quantity,
            stock_available=product["quantity"],
        )

    def test_cash_checkout_persists_every_transaction_record(self):
        product = self.add_product(price=40, quantity=10)
        sale = self.service.checkout(
            user_id=self.admin.id,
            cart_items=[self.cart_row(product, 2)],
            payment_method="CASH",
            received_amount=100,
            discount=5,
        )
        self.assertEqual(sale["subtotal"], 80)
        self.assertEqual(sale["discount"], 5)
        self.assertEqual(sale["total"], 75)
        self.assertEqual(sale["change_due"], 25)

        self.assertEqual(
            self.conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0], 1
        )
        self.assertEqual(
            self.conn.execute("SELECT COUNT(*) FROM sale_items").fetchone()[0], 1
        )
        payment = self.conn.execute("SELECT * FROM payments").fetchone()
        self.assertEqual(payment["method"], "CASH")
        self.assertEqual(payment["amount"], 75)
        movement = self.conn.execute("SELECT * FROM stock_movements").fetchone()
        self.assertEqual(movement["qty"], -2)
        self.assertEqual(movement["movement_type"], "OUT")
        self.assertEqual(self.service.get_product_by_sku("SKU-001")["quantity"], 8)
        self.assertEqual(
            self.conn.execute("SELECT action FROM audit_logs").fetchone()["action"],
            "SALE",
        )

        sales_file = self.temp_path / "sales.xlsx"
        self.assertTrue(sales_file.exists())
        workbook = load_workbook(sales_file, data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(2, 1).value, sale["sale_id"])

    def test_checkout_validation_and_transaction_rollback(self):
        product = self.add_product(price=50, quantity=2)
        row = self.cart_row(product, 1)
        with self.assertRaisesRegex(ValueError, "Cart is empty"):
            self.service.checkout(self.admin.id, [], "CASH", 100)
        with self.assertRaisesRegex(ValueError, "negative"):
            self.service.checkout(self.admin.id, [row], "CASH", 100, discount=-1)
        with self.assertRaisesRegex(ValueError, "Insufficient"):
            self.service.checkout(self.admin.id, [row], "CASH", 1)

        overstocked = self.cart_row(product, 5)
        with self.assertRaisesRegex(ValueError, "Not enough stock"):
            self.service.checkout(self.admin.id, [overstocked], "CASH", 500)
        self.assertEqual(self.conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0], 0)
        self.assertEqual(self.service.get_product_by_sku("SKU-001")["quantity"], 2)

    def test_sales_summary_and_top_skus(self):
        first = self.add_product(sku="FIRST", barcode="B1", price=10, quantity=10)
        second = self.add_product(sku="SECOND", barcode="B2", price=20, quantity=10)
        self.service.checkout(
            self.admin.id, [self.cart_row(first, 3)], "CARD", 30
        )
        self.service.checkout(
            self.admin.id, [self.cart_row(second, 1)], "E-WALLET", 20
        )
        summary = self.service.get_sales_summary()
        self.assertEqual(summary["totals"]["today"]["count"], 2)
        self.assertEqual(summary["totals"]["today"]["total"], 50)
        self.assertEqual(summary["top_skus"][0]["sku"], "FIRST")
        self.assertEqual(summary["top_skus"][0]["quantity"], 3)


class SpreadsheetTests(DatabaseTestCase):
    HEADERS = ["Name", "SKU", "Category", "Price", "Quantity", "Brand", "Barcode"]

    def test_csv_reader_normalizes_headers_and_rows(self):
        csv_path = self.temp_path / "products.csv"
        csv_path.write_text(
            "Name,SKU,Category,Price,Quantity,Brand,Barcode\n"
            "Tea,TEA-1,Drinks,35.5,4,House,490001\n",
            encoding="utf-8",
        )
        rows, headers = importers._read_product_rows(csv_path, None)
        self.assertTrue(importers._has_required_headers(headers))
        self.assertEqual(rows[0]["name"], "Tea")
        self.assertEqual(rows[0]["_row_number"], 2)

    def test_excel_reader_and_product_export(self):
        input_path = self.temp_path / "products.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(self.HEADERS)
        sheet.append(["Rice", "RICE-1", "Food", 55, 8, "Local", "490002"])
        workbook.save(input_path)

        rows, headers = importers._read_product_rows(input_path, None)
        self.assertTrue(importers._has_required_headers(headers))
        self.assertEqual(rows[0]["sku"], "RICE-1")
        self.service.import_products(rows)

        output_path = self.temp_path / "export.xlsx"
        result = importers._export_products_worker(self.service, output_path)
        self.assertTrue(result["success"])
        self.assertEqual(result["count"], 1)
        exported = load_workbook(output_path, data_only=True).active
        self.assertEqual(exported.cell(1, 1).value, "Name")
        self.assertEqual(exported.cell(2, 2).value, "RICE-1")

    def test_invalid_spreadsheet_headers_are_rejected(self):
        csv_path = self.temp_path / "invalid.csv"
        csv_path.write_text("Name,SKU\nIncomplete,ONE\n", encoding="utf-8")
        with patch("app.ui.importers.messagebox.showerror") as showerror:
            rows, headers = importers._read_product_rows(csv_path, None)
        self.assertIsNone(rows)
        self.assertIsNone(headers)
        showerror.assert_called_once()


class ControllerTests(DatabaseTestCase):
    class FakeRoot:
        def __init__(self):
            self.attribute_calls = []
            self.destroyed = False

        def attributes(self, *args):
            self.attribute_calls.append(args)

        def destroy(self):
            self.destroyed = True

    class FakeLoginView:
        def __init__(self):
            self.visible = False
            self.cleared = False

        def show(self):
            self.visible = True

        def hide(self):
            self.visible = False

        def clear_form(self):
            self.cleared = True

    class FakePosView:
        def __init__(self):
            self.visible = False
            self.label = None
            self.cart_rows = []
            self.focused = False
            self.entry_values = ("", "1")
            self.reset = False

        def show(self):
            self.visible = True

        def hide(self):
            self.visible = False

        def set_user_label(self, value):
            self.label = value

        def update_cart(self, rows):
            self.cart_rows = list(rows)

        def focus_sku(self):
            self.focused = True

        def get_entry_values(self):
            return self.entry_values

        def reset_entry_fields(self):
            self.reset = True

    def make_controller(self):
        root = self.FakeRoot()
        state = MainWindowState(self.temp_path / "products.xlsx", get_palette())
        cart = CartService(state.cart)
        controller = PosController(root, self.service, state, cart)
        login_view = self.FakeLoginView()
        pos_view = self.FakePosView()
        controller.bind_views(login_view, pos_view)
        return controller, root, state, cart, login_view, pos_view

    def test_login_navigation_and_invalid_credentials(self):
        controller, root, state, _, login_view, pos_view = self.make_controller()
        controller.login("admin", "admin")
        self.assertEqual(state.current_user.username, "admin")
        self.assertFalse(login_view.visible)
        self.assertTrue(pos_view.visible)
        self.assertTrue(pos_view.focused)
        self.assertIn("Administrator", pos_view.label)
        self.assertIn(("-fullscreen", True), root.attribute_calls)

        with patch("app.ui.controllers.pos_controller.messagebox.showerror") as error:
            controller.login("admin", "wrong")
        error.assert_called_once()

    def test_controller_adds_product_and_warns_low_stock_once(self):
        product = self.add_product(quantity=3)
        controller, _, state, cart, _, pos_view = self.make_controller()
        state.current_user = self.admin
        pos_view.entry_values = (product["sku"], "1")
        with patch("app.ui.controllers.pos_controller.messagebox.showwarning") as warning:
            controller.add_item_from_entry()
            controller.add_item_from_entry()
        self.assertEqual(warning.call_count, 1)
        self.assertEqual(cart.checkout_items()[0].quantity, 2)
        self.assertTrue(pos_view.reset)

    def test_logout_resets_session_without_touching_database(self):
        controller, _, state, cart, login_view, _ = self.make_controller()
        state.current_user = self.admin
        state.cart[1] = CartRow(1, "Item", "SKU", 1, 1, 1)
        with patch.object(controller, "_should_backup", return_value=False):
            controller.logout()
        self.assertIsNone(state.current_user)
        self.assertFalse(cart.has_items())
        self.assertTrue(login_view.visible)
        self.assertTrue(login_view.cleared)


class UtilityAndResourceTests(unittest.TestCase):
    def test_currency_format_state_reset_and_icon_resources(self):
        self.assertEqual(format_currency(1234.5), "P1,234.50")
        state = MainWindowState(Path("products.xlsx"), get_palette())
        state.current_user = object()
        state.cart[1] = CartRow(1, "Item", "SKU", 1, 1, 1)
        state.low_stock_notified.add(1)
        state.reset_session()
        self.assertIsNone(state.current_user)
        self.assertEqual(state.cart, {})
        self.assertEqual(state.low_stock_notified, set())

        self.assertTrue(APP_ICON_PATH.exists())
        self.assertEqual(APP_ICON_PATH.read_bytes()[:4], b"\x00\x00\x01\x00")
        login_asset = APP_ICON_PATH.with_name("pos_login.png")
        self.assertTrue(login_asset.exists())
        self.assertEqual(login_asset.read_bytes()[:8], b"\x89PNG\r\n\x1a\n")


class TkinterInterfaceTests(DatabaseTestCase):
    """Construct every custom screen when a usable display is available."""

    def test_all_views_and_dialogs_construct(self):
        try:
            root = tk.Tk()
        except tk.TclError as exc:
            self.skipTest(f"Tkinter display/runtime unavailable: {exc}")

        root.withdraw()
        try:
            apply_theme(ttk.Style(root), get_palette())
            apply_window_icon(root)
            container = ttk.Frame(root)
            container.pack(fill="both", expand=True)
            noop = lambda *args, **kwargs: None

            login = LoginView(container, noop)
            pos = PosView(
                container,
                on_logout=noop,
                on_load_products=noop,
                on_export_products=noop,
                on_restore_db=noop,
                on_show_reports=noop,
                on_manage_users=noop,
                on_add_item=noop,
                on_remove_item=noop,
                on_clear_cart=noop,
                on_checkout=noop,
            )
            login.show()
            root.update_idletasks()
            self.assertTrue(login.username_entry.winfo_exists())
            login.hide()
            pos.show()
            root.update_idletasks()
            self.assertTrue(pos.cart_tree.winfo_exists())

            dialogs = []
            dialogs.append(AdminAuthDialog(root))
            dialogs.append(PaymentDialog(root, subtotal=100))
            dialogs.append(
                ReportDialog(
                    root,
                    {
                        "totals": {
                            "today": {"total": 0, "count": 0},
                            "week": {"total": 0, "count": 0},
                            "month": {"total": 0, "count": 0},
                        },
                        "top_skus": [],
                    },
                )
            )
            backup = self.temp_path / "pos_backup_20260101_000000_000000.db"
            backup.touch()
            dialogs.append(RestoreBackupDialog(root, [backup]))
            dialogs.append(
                ManageUsersDialog(
                    root,
                    self.service.list_users(),
                    on_update_username=noop,
                    on_reset_pin=noop,
                    on_create_user=noop,
                    on_set_active=noop,
                )
            )
            dialogs.append(ErrorListDialog(root, "Errors", ["Example error"]))
            dialogs.append(ProgressDialog(root, "Progress", 10))
            dialogs.append(
                ImportPreviewDialog(
                    root,
                    columns=[("sku", "SKU"), ("name", "Name")],
                    rows=[{"sku": "SKU", "name": "Product"}],
                    summary={"total": 1, "valid": 1, "invalid": 0, "add": 1, "update": 0},
                    errors=[],
                )
            )
            root.update_idletasks()
            for dialog in reversed(dialogs):
                self.assertTrue(dialog.top.winfo_exists())
                dialog.top.destroy()
        finally:
            root.destroy()


if __name__ == "__main__":
    unittest.main(verbosity=2)
